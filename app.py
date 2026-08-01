import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# =========================================================
# STREAMLIT PAGE CONFIG & MODERN LIGHT UI STYLING
# =========================================================
st.set_page_config(
    page_title="Fleet Mileage Analytics & Report Generator",
    page_icon="🚛",
    layout="wide"
)

# Executive Light Theme Styling
st.markdown("""
    <style>
    .main {
        background-color: #F8FAFC;
        color: #0F172A;
    }
    .stApp {
        background-color: #F8FAFC;
    }
    .css-1d39121, .stSidebar {
        background-color: #FFFFFF !important;
        border-right: 1px solid #E2E8F0;
    }
    h1, h2, h3 {
        color: #0F172A !important;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 700;
    }
    .stButton>button {
        background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
        color: white;
        font-weight: 600;
        border: none;
        border-radius: 8px;
        padding: 10px 24px;
        box-shadow: 0 4px 6px -1px rgba(30, 58, 138, 0.2);
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background: linear-gradient(135deg, #1E40AF 0%, #1D4ED8 100%);
        box-shadow: 0 6px 12px -1px rgba(30, 58, 138, 0.3);
    }
    </style>
""", unsafe_allow_html=True)

# =========================================================
# HELPER FUNCTIONS & STRICT DATA PARSING
# =========================================================

def parse_period_to_seconds(period_str):
    """Converts HH:MM:SS or HH:MM string to total seconds."""
    try:
        parts = str(period_str).strip().split(':')
        if len(parts) == 3:
            return int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0])*3600 + int(parts[1])*60
    except:
        return 0
    return 0

def format_seconds_to_hhmm(seconds):
    """Formats total seconds into HH:MM string."""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"

def clean_and_parse_input(uploaded_file):
    """
    STRICT DATA PARSER & MASTER LOOKUP
    Extracts complete vehicle metadata and cleans every individual row without date loss.
    """
    df_raw = pd.read_excel(uploaded_file)
    
    if '#' in str(df_raw.columns[0]) or 'Device' in str(df_raw.columns[1]):
        df = df_raw.copy()
    else:
        headers = [str(val).replace('\xa0', '').strip() for val in df_raw.iloc[0].values]
        df = df_raw.iloc[1:].copy()
        df.columns = headers

    # Clean whitespace & \xa0
    for col in df.columns:
        df[col] = df[col].astype(str).str.replace('\xa0', '').str.strip()

    # Create master lookup dictionary for vehicles
    device_map = {}
    reg_map = {}

    invalid_vals = ['', 'nan', 'None', '-']

    for _, row in df.iterrows():
        dev = row.get('Device', '').strip()
        reg = row.get('Reg#', '').strip()
        grp = row.get('Group', '').strip()
        vtype = row.get('Vehicle Type', '').strip()
        purp = row.get('Purpose', '').strip()
        th_k = row.get('TH Kms', '').strip()
        th_t = row.get('TH Tme', '').strip()

        if dev not in invalid_vals and reg not in invalid_vals:
            info = {
                'Device': dev, 'Reg#': reg, 'Group': grp,
                'Vehicle Type': vtype, 'Purpose': purp,
                'TH Kms': th_k, 'TH Tme': th_t
            }
            device_map[dev] = info
            reg_map[reg.upper()] = info

    # Fill missing metadata while preserving row integrity
    for idx in df.index:
        dev = df.at[idx, 'Device']
        reg = df.at[idx, 'Reg#']

        ref = None
        if dev in device_map:
            ref = device_map[dev]
        elif reg.upper() in reg_map:
            ref = reg_map[reg.upper()]

        if ref:
            for col in ['Device', 'Reg#', 'Group', 'Vehicle Type', 'Purpose', 'TH Kms', 'TH Tme']:
                curr = df.at[idx, col]
                if curr in invalid_vals or pd.isna(curr):
                    df.at[idx, col] = ref.get(col, '')

    # Convert distance & time
    df['Distance_num'] = pd.to_numeric(df['Distance'], errors='coerce').fillna(0.0)
    df['Date_parsed'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Period_seconds'] = df['Period'].apply(parse_period_to_seconds)
    
    return df

def apply_cell_locking_footer(ws, last_row, dev_name="Muhammad Ashaan", rep_name="Ahmad Raza"):
    """Adds fixed locked cells for Developer and Reporter."""
    foot_row = last_row + 2
    
    c_dev = ws.cell(row=foot_row, column=1, value=f"Developed By: {dev_name}")
    c_dev.font = Font(name='Segoe UI', size=9, bold=True, color='475569')
    c_dev.alignment = Alignment(horizontal='left', vertical='center')
    c_dev.protection = openpyxl.styles.Protection(locked=True)
    
    c_rep = ws.cell(row=foot_row, column=3, value=f"Report By: {rep_name}")
    c_rep.font = Font(name='Segoe UI', size=9, bold=True, color='475569')
    c_rep.alignment = Alignment(horizontal='left', vertical='center')
    c_rep.protection = openpyxl.styles.Protection(locked=True)


def build_sunday_report_excel(df, city_name, month_name):
    """Generates Sunday Working Vehicles Excel Sheet."""
    # Sunday filters distance > 0
    df_valid = df[
        (df['Period'] != '00:00:00') & 
        (df['Period'] != '0:00:00') & 
        (df['Distance_num'] > 0)
    ].copy()

    sundays = sorted(df_valid[df_valid['Date_parsed'].dt.dayofweek == 6]['Date_parsed'].dropna().unique())
    dates_to_process = sundays if len(sundays) > 0 else sorted(df_valid['Date_parsed'].dropna().unique())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for idx, dt in enumerate(dates_to_process, 1):
        dt_obj = pd.to_datetime(dt)
        dt_str = dt_obj.strftime('%d-%b-%Y') if pd.notnull(dt_obj) else "N/A"
        sheet_title = f"Sunday {idx} ({dt_obj.strftime('%b %d')})" if len(sundays) > 0 and pd.notnull(dt_obj) else f"Day {idx}"
        
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.views.sheetView[0].showGridLines = True

        df_day = df_valid[df_valid['Date_parsed'] == dt].sort_values(by='Distance_num', ascending=False)

        # Title Header
        ws.merge_cells('A1:K1')
        t_cell = ws['A1']
        t_cell.value = f"SUNDAY WORKING VEHICLES MILEAGE REPORT — {city_name.upper()}"
        t_cell.font = Font(name='Segoe UI', size=13, bold=True, color='FFFFFF')
        t_cell.fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        t_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 38

        # Subtitle
        ws.merge_cells('A2:K2')
        s_cell = ws['A2']
        s_cell.value = f"Date: {dt_str}   |   Month: {month_name}   |   Active Fleet: {len(df_day)} Vehicles"
        s_cell.font = Font(name='Segoe UI', size=9.5, italic=True, color='0284C7')
        s_cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        s_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        headers = ['Sr #', 'Device ID', 'Reg #', 'Group / Region', 'Vehicle Type', 'Purpose', 'TH Kms', 'TH Time', 'Date', 'Distance (Km)', 'Period (HH:MM)']
        ws.row_dimensions[4].height = 26
        
        h_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        h_font = Font(name='Segoe UI', size=9.5, bold=True, color='FFFFFF')

        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.fill = h_fill
            c.font = h_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border

        start_row = 5
        total_sec = 0
        for r_idx, (_, row) in enumerate(df_day.iterrows(), start_row):
            ws.row_dimensions[r_idx].height = 20
            bg_color = 'F8FAFC' if r_idx % 2 == 0 else 'FFFFFF'
            r_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

            total_sec += row['Period_seconds']

            d_vals = [
                r_idx - start_row + 1,
                row['Device'],
                row['Reg#'],
                row['Group'],
                row['Vehicle Type'],
                row['Purpose'],
                float(row['TH Kms']) if str(row['TH Kms']).replace('.','',1).isdigit() else row['TH Kms'],
                row['TH Tme'],
                row['Date'],
                float(row['Distance_num']),
                row['Period']
            ]

            for c_idx, val in enumerate(d_vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = r_fill
                cell.border = thin_border
                cell.font = Font(name='Segoe UI', size=9)

                if c_idx in [1, 2, 3, 7, 8, 9, 11]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx == 10:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Total Row
        tot_row = start_row + len(df_day)
        ws.row_dimensions[tot_row].height = 24
        tot_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        tot_font = Font(name='Segoe UI', size=9.5, bold=True, color='0F172A')

        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=9)
        tot_label = ws.cell(row=tot_row, column=1, value="TOTAL SUNDAY WORKING SUMMARY")
        tot_label.font = tot_font
        tot_label.alignment = Alignment(horizontal='right', vertical='center')

        for c in range(1, 10):
            ws.cell(row=tot_row, column=c).border = thin_border
            ws.cell(row=tot_row, column=c).fill = tot_fill

        tot_val = ws.cell(row=tot_row, column=10, value=f"=SUM(J{start_row}:J{tot_row-1})" if len(df_day)>0 else 0)
        tot_val.font = tot_font
        tot_val.fill = tot_fill
        tot_val.alignment = Alignment(horizontal='right', vertical='center')
        tot_val.number_format = '#,##0.00'
        tot_val.border = thin_border

        tot_p = ws.cell(row=tot_row, column=11, value=format_seconds_to_hhmm(total_sec))
        tot_p.font = tot_font
        tot_p.fill = tot_fill
        tot_p.alignment = Alignment(horizontal='center', vertical='center')
        tot_p.border = thin_border

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == 'A':
                ws.column_dimensions[col_letter].width = 12  # Exact width 12 for Sr#
            else:
                max_l = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col_letter].width = max(max_l + 3, 14)

        apply_cell_locking_footer(ws, tot_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def build_evening_report_excel(df, city_name, month_name, target_vehicles):
    """
    EVENING SHIFT REPORT
    Includes 0 time period entries and complete monthly dates for target vehicles.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evening Shift Report"
    ws.views.sheetView[0].showGridLines = True

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Main Header Block
    ws.merge_cells('A1:K1')
    t_cell = ws['A1']
    t_cell.value = f"EVENING VEHICLES SHIFT MILEAGE REPORT — {city_name.upper()}"
    t_cell.font = Font(name='Segoe UI', size=13, bold=True, color='FFFFFF')
    t_cell.fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    t_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:K2')
    s_cell = ws['A2']
    s_cell.value = f"Month Period: {month_name}   |   Target Evening Fleet: {', '.join([v.upper() for v in target_vehicles])}"
    s_cell.font = Font(name='Segoe UI', size=9.5, italic=True, color='0284C7')
    s_cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    s_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    current_row = 4
    headers = ['Sr #', 'Device ID', 'Reg #', 'Group / Region', 'Vehicle Type', 'Purpose', 'TH Kms', 'TH Time', 'Date', 'Distance (Km)', 'Period (HH:MM)']

    for veh in target_vehicles:
        # DO NOT exclude 0 time period or 0 distance entries for evening report!
        df_veh = df[df['Reg#'].str.upper() == veh.upper()].copy()
        
        # Sort by Date if available, else Distance
        if 'Date_parsed' in df_veh.columns and df_veh['Date_parsed'].notnull().any():
            df_veh = df_veh.sort_values(by='Date_parsed', ascending=True)

        # Vehicle Section Banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        v_cell = ws.cell(row=current_row, column=1, value=f"VEHICLE REGISTRATION: {veh.upper()}   (Total Monthly Entries: {len(df_veh)})")
        v_cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        v_cell.fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        v_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Table Column Headers
        ws.row_dimensions[current_row].height = 24
        h_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        h_font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')

        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=current_row, column=col_idx, value=h)
            c.fill = h_fill
            c.font = h_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border

        current_row += 1

        if len(df_veh) == 0:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            empty_info = ws.cell(row=current_row, column=1, value="No records found for this vehicle in the uploaded sheet.")
            empty_info.font = Font(name='Segoe UI', size=9, italic=True, color='64748B')
            empty_info.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(1, 12):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        else:
            start_v_row = current_row
            v_sec = 0
            for r_idx, (_, row) in enumerate(df_veh.iterrows(), start_v_row):
                ws.row_dimensions[r_idx].height = 20
                bg_color = 'F8FAFC' if r_idx % 2 == 0 else 'FFFFFF'
                r_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                v_sec += row['Period_seconds']

                d_vals = [
                    r_idx - start_v_row + 1,
                    row['Device'],
                    row['Reg#'],
                    row['Group'],
                    row['Vehicle Type'],
                    row['Purpose'],
                    float(row['TH Kms']) if str(row['TH Kms']).replace('.','',1).isdigit() else row['TH Kms'],
                    row['TH Tme'],
                    row['Date'],
                    float(row['Distance_num']),
                    row['Period']
                ]

                for c_idx, val in enumerate(d_vals, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.fill = r_fill
                    cell.border = thin_border
                    cell.font = Font(name='Segoe UI', size=9)

                    if c_idx in [1, 2, 3, 7, 8, 9, 11]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif c_idx == 10:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                current_row += 1

            # Vehicle Total Row
            tot_row = current_row
            ws.row_dimensions[tot_row].height = 23
            tot_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
            tot_font = Font(name='Segoe UI', size=9, bold=True, color='0F172A')

            ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=9)
            tot_label = ws.cell(row=tot_row, column=1, value=f"TOTAL MONTHLY MILEAGE FOR {veh.upper()}")
            tot_label.font = tot_font
            tot_label.alignment = Alignment(horizontal='right', vertical='center')

            for c in range(1, 10):
                ws.cell(row=tot_row, column=c).border = thin_border
                ws.cell(row=tot_row, column=c).fill = tot_fill

            tot_val = ws.cell(row=tot_row, column=10, value=f"=SUM(J{start_v_row}:J{tot_row-1})")
            tot_val.font = tot_font
            tot_val.fill = tot_fill
            tot_val.alignment = Alignment(horizontal='right', vertical='center')
            tot_val.number_format = '#,##0.00'
            tot_val.border = thin_border

            tot_p = ws.cell(row=tot_row, column=11, value=format_seconds_to_hhmm(v_sec))
            tot_p.font = tot_font
            tot_p.fill = tot_fill
            tot_p.alignment = Alignment(horizontal='center', vertical='center')
            tot_p.border = thin_border

            current_row += 1

        current_row += 2  # Spacing between vehicles

    # Set Column Widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = 12  # Exact 12 Width
        else:
            max_l = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_l + 3, 14)

    # Footer Locking
    apply_cell_locking_footer(ws, current_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =========================================================
# UI DASHBOARD
# =========================================================

st.title("🚛 Fleet Mileage Report Generator")
st.markdown("Aesthetic & Professional Excel Report Processing Platform.")

st.divider()

col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("⚙️ Control Panel")
    
    uploaded_file = st.file_uploader(
        "Upload Monthly Raw Mileage Excel (.xlsx)",
        type=["xlsx", "xls"]
    )
    
    tehsil_choice = st.selectbox(
        "Select Tehsil / City Name",
        ["Kamoke", "Nowshera Virkan", "Custom"]
    )
    
    city_name = st.text_input("Tehsil Name", value="Kamoke") if tehsil_choice == "Custom" else tehsil_choice

    month_name = st.selectbox(
        "Select Report Month",
        ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"],
        index=6
    )
    
    if city_name.lower() == "kamoke":
        default_vehs = "GBA-915, GBA-917"
    elif city_name.lower() in ["nowshera virkan", "nowshera"]:
        default_vehs = "RIC-159, RIC-165, TT-240"
    else:
        default_vehs = "GBA-915, GBA-917"

    veh_input = st.text_input("Target Evening Vehicles (Comma Separated)", value=default_vehs)
    target_vehicles = [v.strip() for v in veh_input.split(",") if v.strip()]

with col2:
    st.subheader("📊 Report Preview & Generation")
    
    if uploaded_file is not None:
        try:
            df_clean = clean_and_parse_input(uploaded_file)
            
            st.success("✅ Master Vehicle Mapping Applied & Data Parsed Successfully!")
            
            m1, m2, m3 = st.columns(3)
            m1.metric("Total Monthly Records", len(df_clean))
            m2.metric("Active Fleet Vehicles", df_clean['Reg#'].nunique())
            m3.metric("Evening Vehicles Target", len(target_vehicles))

            st.markdown("##### Cleaned Data Verification (With Dates Intact)")
            st.dataframe(df_clean[['Device', 'Reg#', 'Group', 'Vehicle Type', 'Date', 'Distance', 'Period']].head(10), use_container_width=True)
            
            st.divider()
            
            st.markdown("### 📥 Generated Reports")
            
            c_btn1, c_btn2 = st.columns(2)
            
            # 1. Sunday Working Report
            s_filename = f"Mileage Report {month_name} All Sunday Working vehicles report Tehsil {city_name}.xlsx"
            s_bytes = build_sunday_report_excel(df_clean, city_name, month_name)
            
            with c_btn1:
                st.download_button(
                    label="📊 Download Sunday Working Report",
                    data=s_bytes,
                    file_name=s_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            # 2. Evening Shift Report
            e_filename = f"Mileage report {month_name} evening vehicles shift report tehsil {city_name}.xlsx"
            e_bytes = build_evening_report_excel(df_clean, city_name, month_name, target_vehicles)
            
            with c_btn2:
                st.download_button(
                    label="🌙 Download Evening Shift Report",
                    data=e_bytes,
                    file_name=e_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
    else:
        st.info("👈 Please upload your raw Excel file to generate error-free reports.")
